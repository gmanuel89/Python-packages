#####
# Author: Manuel Galli
# e-mail: gmanuel89@gmail.com / manuel.galli@revvity.com
# Updated date: 2022-10-07
#####

## Import libraries and functions
import openpyxl
from openpyxl.utils import get_column_letter

## Read spreadsheet from workbook
def read_workbook(input_workbook_file: str, selected_spreadsheet=None, output_format='list') -> list[list] | list[dict]:
    # Initialise output variable
    input_file_lines = []
    # Read the workbook
    workbook_content = openpyxl.load_workbook(input_workbook_file)
    # Select the sheet
    if selected_spreadsheet is None or selected_spreadsheet == '':
        spreadsheet_sheet = workbook_content.active
    else:
        if selected_spreadsheet in workbook_content.sheetnames:
            spreadsheet_sheet = workbook_content[selected_spreadsheet]
        else:
            spreadsheet_sheet = workbook_content.active
    # Retrieve the sheet content
    spreadsheet_sheet_content = []
    for r in range(len(list(spreadsheet_sheet.rows))-1):
        row_values = []
        # Add the column values to the row list
        for c in range(spreadsheet_sheet.max_column):
            row_values.append(spreadsheet_sheet[get_column_letter(c+1)+str(r+1)].value)
        spreadsheet_sheet_content.append(row_values)
    # Leave it as list or convert it to dictionary
    if output_format == 'dictionary':
        converted_csv_content = []
        # Header
        csv_header = spreadsheet_sheet_content[0]
        # Remove it from the content
        spreadsheet_sheet_content.remove(csv_header)
        # For each row...
        for row in spreadsheet_sheet_content:
            row_dict = {}
            # Add the cell value of the corresponding column in the dictionary (with the key being the column name)
            for c in range(len(row)):
                row_dict[csv_header[c]] = row[c]
            # Add the row dictionary to the list of row dictionaries
            converted_csv_content.append(row_dict)
        # Output
        input_file_lines = converted_csv_content
    else:
        input_file_lines = spreadsheet_sheet_content
    # return
    return input_file_lines