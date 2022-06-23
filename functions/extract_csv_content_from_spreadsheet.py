# Import libraries
import openpyxl
from openpyxl.utils import get_column_letter
from functions.get_spreadsheet_header import *
from functions.convert_csv_content import *
from openpyxl.worksheet.worksheet import Worksheet

## Extract the content of a openpyxl Worksheet to a list (output type can be 'list' or 'dictionary')
def extract_csv_content_from_spreadsheet(spreadsheet_sheet: Worksheet, output_type='list', column_indices=None) -> list[list] | list[dict]:
    # Initialise output variable
    spreadsheet_sheet_content = []
    # Scroll all the rows
    for r in range(len(list(spreadsheet_sheet.rows))-1):
        row_values = []
        # Add the column values to the row list
        for c in range(spreadsheet_sheet.max_column):
            # Do not add it if not of interest (only if the list exists)
            if column_indices and c+1 in column_indices:
                row_values.append(spreadsheet_sheet[get_column_letter(c+1)+str(r+1)].value)
        spreadsheet_sheet_content.append(row_values)
    # Leave it as list or convert it to dictionary
    if output_type == 'dictionary':
        spreadsheet_sheet_content = convert_csv_content(spreadsheet_sheet_content)
    else:
        spreadsheet_sheet_content = spreadsheet_sheet_content 
    # return
    return spreadsheet_sheet_content
