#####
# Author: Manuel Galli
# e-mail: gmanuel89@gmail.com / manuel.galli@perkinelmer.com
# Updated date: 2022-10-07
#####

## Import libraries
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

## Get the header of a Spreadsheet
def get_spreadsheet_header(spreadsheet: Worksheet) -> list[str]:
    # Initialise output
    spreadsheet_header = []
    # Get the header value
    for c in range(spreadsheet.max_column):
        cell_value = spreadsheet[get_column_letter(c+1)+'1'].value
        spreadsheet_header.append(cell_value)
    # return
    return spreadsheet_header
