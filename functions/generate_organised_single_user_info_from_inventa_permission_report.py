## Import libraries and functions
import openpyxl
from openpyxl.utils import get_column_letter
from functions.excel_workbook_handling import *
from functions.common import *

## Generate single user info report from Inventa Permission Report
def generate_organised_single_user_info_from_inventa_permission_report(signals_inventa_user_security_report_content: openpyxl.Workbook, selected_user: str) -> dict:
    # Initialise output
    single_user_info = {}
    # Login Name
    single_user_info['Login Name'] = selected_user
    # Name
    project_user_access_sheet = signals_inventa_user_security_report_content['Project User Access'] # "Project User Access" sheet (all users are here)
    for r in range(len(list(project_user_access_sheet.rows))):
        if project_user_access_sheet[get_column_letter(2)+str(r+1)].value == selected_user:
            single_user_info['Name'] = project_user_access_sheet[get_column_letter(1)+str(r+1)].value
            break
    # User Groups
    user_groups_assignments_sheet = signals_inventa_user_security_report_content['User Group Assignments']
    user_groups_assignments_sheet_content = extract_csv_content_from_spreadsheet(user_groups_assignments_sheet, output_type='list', column_indices=[2,3])
    key_values = []
    values_column = []
    user_groups_assignments_sheet_content.remove(user_groups_assignments_sheet_content[0])
    for row in user_groups_assignments_sheet_content:
        key_values.append(row[0])
        values_column.append(row[1])
    single_user_info['Groups'] = generate_string_with_concatenated_key_values(key_values, values_column, ' | ', selected_key_values=[selected_user])
    single_user_info['Groups'] = single_user_info['Groups'].get(selected_user)
    # Projects and Project permissions
    project_user_access_sheet = signals_inventa_user_security_report_content['Project User Access']
    project_user_access_sheet_content = extract_csv_content_from_spreadsheet(project_user_access_sheet, output_type='list', column_indices=[2,3,4])
    project_user_access_sheet_content.remove(project_user_access_sheet_content[0])
    key_values = []
    values_column = []
    for row in project_user_access_sheet_content: # projects
        key_values.append(row[0])
        values_column.append(row[1])
    single_user_info['Projects'] = generate_string_with_concatenated_key_values(key_values, values_column, ' | ', selected_key_values=[selected_user])
    single_user_info['Projects'] = single_user_info['Projects'].get(selected_user)
    key_values = []
    values_column = []
    for row in project_user_access_sheet_content: # project permissions
        key_values.append(row[0])
        values_column.append(row[2])
    single_user_info['Project Permissions'] = generate_string_with_concatenated_key_values(key_values, values_column, ' | ', selected_key_values=[selected_user])
    single_user_info['Project Permissions'] = single_user_info['Project Permissions'].get(selected_user)
    # Measurement Types and MType permissions
    mtype_user_access_sheet = signals_inventa_user_security_report_content['Measurement Type User Access']
    mtype_user_access_sheet_content = extract_csv_content_from_spreadsheet(mtype_user_access_sheet, output_type='list', column_indices=[2,4,5])
    mtype_user_access_sheet_content.remove(mtype_user_access_sheet_content[0])
    key_values = []
    values_column = []
    for row in mtype_user_access_sheet_content: # mtypes
        key_values.append(row[0])
        values_column.append(row[1])
    single_user_info['Measurement Types'] = generate_string_with_concatenated_key_values(key_values, values_column, ' | ', selected_key_values=[selected_user])
    single_user_info['Measurement Types'] = single_user_info['Measurement Types'].get(selected_user)
    key_values = []
    values_column = []
    for row in mtype_user_access_sheet_content: # mtype permissions
        key_values.append(row[0])
        if str(row[2]).lower() == 'true':
            row[2] = 'View'
        else:
            row[2] = 'No view'
        values_column.append(row[2])
    single_user_info['MType Permissions'] = generate_string_with_concatenated_key_values(key_values, values_column, ' | ', selected_key_values=[selected_user])
    single_user_info['MType Permissions'] = single_user_info['MType Permissions'].get(selected_user)
    # Administrator
    project_user_access_sheet = signals_inventa_user_security_report_content['Project User Access'] # "Project User Access" sheet (all users are here)
    for r in range(len(list(project_user_access_sheet.rows))):
        if project_user_access_sheet[get_column_letter(2)+str(r+1)].value == selected_user:
            single_user_info['Administrator'] = project_user_access_sheet[get_column_letter(5)+str(r+1)].value
            break
    # return
    return single_user_info

## TABLE
# Login Name, Name, Groups, Projects, Project Permissions, Measurement Types, MType Permissions, Administrator 